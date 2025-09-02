import yfinance as yf
from openpyxl import load_workbook
import requests
import json
import datetime

url = "https://umihico.github.io/kabu-json-all-stock-list/all_stocks.json"
response = requests.get(url)
stucks = json.loads(response.text)

wb = load_workbook("/Users/sugiyamawataru/Documents/stuck_test.xlsx")
ws = wb.active
ws["A1"] = "銘柄コード"
ws["B1"] = "銘柄名"
ws["C1"] = "EPS"
ws["D1"] = "PER"
ws["E1"] = "2年成長率"
ws["F1"] = "5年成長率"

ws["H1"] = "銘柄コード"
ws["I1"] = "銘柄名"
ws["J1"] = "EPS"
ws["K1"] = "PER"
ws["L1"] = "3ヶ月成長率"
ws["M1"] = "6ヶ月成長率"


url = "https://umihico.github.io/kabu-json-all-stock-list/all_stocks.json"
response = requests.get(url)
tickers = json.loads(response.text)
long_row = 2
medium_row = 2


                  
print()

for ticker in tickers:
    stock = yf.Ticker(ticker["コード"] + ".T")

    info = stock.info
    today_data = stock.history(period="1d") #今日の株価
    end = datetime.datetime.today() #成長率のためのデータ
    #長期
    start_for_2years = end - datetime.timedelta(days = 730)
    start_for_5years = end - datetime.timedelta(days = 1825)
    growth_data_for_2years = stock.history(start=start_for_2years.strftime("%Y-%m-%d"), end=end.strftime("%Y-%m-%d"))
    growth_data_for_5years = stock.history(start=start_for_5years.strftime("%Y-%m-%d"), end=end.strftime("%Y-%m-%d"))

    #中期
    start_for_3monthes = end - datetime.timedelta(days = 90)
    start_for_6monthes = end - datetime.timedelta(days = 180)
    growth_data_for_3monthes = stock.history(start=start_for_3monthes.strftime("%Y-%m-%d"), end=end.strftime("%Y-%m-%d"))
    growth_data_for_6monthes = stock.history(start=start_for_6monthes.strftime("%Y-%m-%d"), end=end.strftime("%Y-%m-%d"))

    if len(today_data["Close"]) != 0: #株価,eps,per,growthの入手
        today_close_data = today_data["Close"][0]
    eps = info.get("trailingEsp", "N/A")
    per = info.get("trailingPE", "N/A")


    if len(growth_data_for_2years) != 0:#2年の成長率
        start_price = growth_data_for_2years["Close"].iloc[0]
        end_price = growth_data_for_2years["Close"].iloc[-1]
        growth_for_2years = (end_price / start_price) -1
    
    if len(growth_data_for_5years) != 0:#5年の成長率
        start_price = growth_data_for_5years["Close"].iloc[0]
        end_price = growth_data_for_5years["Close"].iloc[-1]
        growth_for_5years = (end_price / start_price) -1

    if len(growth_data_for_3monthes) != 0:#3ヶ月の成長率
        start_price = growth_data_for_3monthes["Close"].iloc[0]
        end_price = growth_data_for_3monthes["Close"].iloc[-1]
        growth_for_3monthes = (end_price / start_price) -1
    
    if len(growth_data_for_6monthes) != 0:#6ヶ月の成長率
        start_price = growth_data_for_6monthes["Close"].iloc[0]
        end_price = growth_data_for_6monthes["Close"].iloc[-1]
        growth_for_6monthes = (end_price / start_price) -1

    if eps == "N/A" and per == "N/A":
        continue
    if (eps == "N/A" and per != "N/A") or (eps != "N/A" and per == "N/A"): #esp,perがかけてる時に補う
        if eps == "N/A" and per !="Infinity":
            eps = today_close_data / per
        else:
            per = today_close_data / eps

  
    if eps > 0 and per < 10 and growth_for_2years >= 0.15:
        ws.cell(row=long_row, column=1, value=ticker["コード"])
        ws.cell(row=long_row, column=2, value=ticker["銘柄名"])
        ws.cell(row=long_row, column=3, value=eps)
        ws.cell(row=long_row, column=4, value=per)
        ws.cell(row=long_row, column=5, value=growth_for_2years)
        ws.cell(row=long_row, column=6, value=growth_for_5years)

        wb.save("/Users/sugiyamawataru/Documents/stuck_test.xlsx")
        long_row += 1
    if eps > 0 and per < 10 and growth_for_3monthes >= 0.1:
        print("コード:",ticker["コード"])
        print("銘柄名:", ticker["銘柄名"])
        """
        ws.cell(row=medium_row, column=8, value=ticker["コード"])
        ws.cell(row=medium_row, column=9, value=ticker["銘柄名"])
        ws.cell(row=medium_row, column=10, value=eps)
        ws.cell(row=medium_row, column=11, value=per)
        ws.cell(row=medium_row, column=12, value=growth_for_2years)
        ws.cell(row=medium_row, column=13, value=growth_for_5years)

        wb.save("/Users/sugiyamawataru/Documents/stuck_test.xlsx")
        medium_row += 1

        
    

wb.save("/Users/sugiyamawataru/Documents/stuck_test.xlsx")
"""